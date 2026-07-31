#!/usr/bin/env python3
"""
LinkedCare 交易级明细导出（电商模板格式）
从 bsk 分批导出的 JSON chunk 文件合并为 CSV + XLSX
表头匹配：工作计划/2026年/电商项目明细导出模板.xlsx

用法：
  python merge_export_template.py [输出目录]
  默认输出目录：~/病例提取
"""

import json, csv, os, sys, glob
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

# ---- 配置 ----
DEFAULT_OUTDIR = os.path.expanduser("~/病例提取")
OUTDIR = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_OUTDIR
os.makedirs(OUTDIR, exist_ok=True)

TS = datetime.now().strftime("%Y%m%d_%H%M")
BASENAME = f"电商项目明细_网电咨询师非空_{TS}"

# 模板表头（与 电商项目明细导出模板.xlsx 一致；【模式B 现仅输出病例ID，不含病历号】）
HEADERS = [
    "患者网电咨询师", "患者姓名", "病例id",
    "收费时间", "现金类实收", "收费机构"
]

# JS 提取对象 → CSV 列 的字段映射顺序
# ⚠️ 病例id = patientId（≠ privateId）。【模式B 现仅输出病例ID，不再输出病历号(privateId)】
KEYS = [
    "onlineConsultantName", "patientName", "patientId",
    "payDateTime", "cashActualReceived", "chargeOrg"
]

# 列宽（字符）
COL_WIDTHS = [16, 12, 18, 20, 14, 30]


def load_chunks(chunk_dir="/tmp", pattern="lc_chunk_*.json"):
    """加载所有分块 JSON 文件，合并为列表"""
    rows = []
    for fpath in sorted(glob.glob(os.path.join(chunk_dir, pattern))):
        with open(fpath, 'r', encoding='utf-8') as f:
            batch = json.load(f)
            if isinstance(batch, list):
                rows.extend(batch)
    return rows


def write_csv(rows, path):
    """写 UTF-8 BOM CSV（Excel 友好）"""
    with open(path, 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.writer(f)
        w.writerow(HEADERS)
        for row in rows:
            w.writerow([row.get(k, '') for k in KEYS])


def write_xlsx(rows, path):
    """写 Excel，套模板样式（加粗表头 + 边框 + 列宽）"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    hdr_font = Font(bold=True)
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 表头行
    for ci, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hdr_font
        c.border = thin
        c.alignment = Alignment(horizontal='center')

    # 数据行
    for ri, row in enumerate(rows, 2):
        for ci, k in enumerate(KEYS, 1):
            v = row.get(k, '')
            # 现金类实收尝试转数字
            if k == 'cashActualReceived' and v != '':
                try:
                    v = float(v)
                except (ValueError, TypeError):
                    pass
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = thin

    # 列宽
    for ci, wdt in enumerate(COL_WIDTHS, 1):
        col_letter = chr(64 + ci) if ci <= 26 else f"A{chr(64 + ci - 26)}"
        ws.column_dimensions[col_letter].width = wdt

    wb.save(path)


def main():
    rows = load_chunks()
    if not rows:
        print("ERROR: 没有找到分块数据文件 (/tmp/lc_chunk_*.json)")
        print("请先运行 extract_detail.js 并用 evaluate 分批导出 __extract2.rows")
        sys.exit(1)

    csv_path = os.path.join(OUTDIR, f"{BASENAME}.csv")
    xlsx_path = os.path.join(OUTDIR, f"{BASENAME}.xlsx")

    write_csv(rows, csv_path)
    print(f"CSV -> {csv_path} ({len(rows)} rows)")

    write_xlsx(rows, xlsx_path)
    print(f"XLSX -> {xlsx_path} ({len(rows)} rows)")

    # 统计
    consultants = set(r.get('onlineConsultantName', '') for r in rows)
    orgs = set(r.get('chargeOrg', '') for r in rows)
    print(f"\n统计: {len(rows)} 行 | {len(consultants)} 位网电咨询师 | {len(orgs)} 个收费机构")


if __name__ == '__main__':
    main()
